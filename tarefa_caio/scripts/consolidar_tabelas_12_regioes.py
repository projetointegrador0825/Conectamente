#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Script consolidado para as Tabelas 12.1 a 12.8 (Aba 1 ou Aba 2).
Caminho atualizado para ambiente Windows conforme image_49c601.png.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def consolidar_tabelas(diretorio_origem=None, aba_indice=0):
    """
    Consolida as Tabelas 12.1 a 12.8 em um arquivo Excel estruturado.
    
    Args:
        diretorio_origem (str): Diretório contendo os arquivos.
        aba_indice (int): 0 = Primeira Aba (Idades), 1 = Segunda Aba (Regiões/UF).
    """
    
    # Atualizado conforme especificação para o diretório de dados no Windows
    if diretorio_origem is None:
        diretorio_origem = r"C:\Users\caiol\Desktop\PROJETO INTEGRADOR 2 ETAPA\projeto-integrador-grupo26\tarefa_caio\dados"
    
    arquivos = [
        ("Tabelas 12.1.xlsx", "12.1.1" if aba_indice == 0 else "12.1.2", "NÃO_TEM_AMIGO_PROXIMO"),
        ("Tabelas 12.2.xlsx", "12.2.1" if aba_indice == 0 else "12.2.2", "MUITO_PREOCUPADO"),
        ("Tabelas 12.3.xlsx", "12.3.1" if aba_indice == 0 else "12.3.2", "TRISTE"),
        ("Tabelas 12.4.xlsx", "12.4.1" if aba_indice == 0 else "12.4.2", "SOZINHO"),
        ("Tabelas 12.5.xlsx", "12.5.1" if aba_indice == 0 else "12.5.2", "PROBLEMA_CONCENTRACAO"),
        ("Tabelas 12.6.xlsx", "12.6.1" if aba_indice == 0 else "12.6.2", "PROBLEMA_INSONIA"),
        ("Tabelas 12.7.xlsx", "12.7.1" if aba_indice == 0 else "12.7.2", "SENTIMENTO_INUTILIDADE"),
        ("Tabelas 12.8.xlsx", "12.8.1" if aba_indice == 0 else "12.8.2", "PENSAMENTO_MORTE"),
    ]
    
    wb_consolidado = Workbook()
    ws_consolidado = wb_consolidado.active
    
    # Definição dinâmica de cabeçalhos e sufixo do arquivo
    if aba_indice == 0:
        cabecalhos = [
            "TABELA_ORIGEM", "NOME_METRICA", "Faixa_Etaria", "Regiao", "rc",
            "Limite_inferior", "Limite_superior",
            "rc_Homem", "Limite_inf_Homem", "Limite_sup_Homem",
            "rc_Mulher", "Limite_inf_Mulher", "Limite_sup_Mulher",
            "rc_Publica", "Limite_inf_Publica", "Limite_sup_Publica",
            "rc_Privada", "Limite_inf_Privada", "Limite_sup_Privada"
        ]
        suffix = "Estruturado_Idades"
    else:
        cabecalhos = [
            "TABELA_ORIGEM", "NOME_METRICA", "Regiao", "Unidade_Federacao", "rc",
            "Limite_inferior", "Limite_superior",
            "rc_Homem", "Limite_inf_Homem", "Limite_sup_Homem",
            "rc_Mulher", "Limite_inf_Mulher", "Limite_sup_Mulher",
            "rc_Publica", "Limite_inf_Publica", "Limite_sup_Publica",
            "rc_Privada", "Limite_inf_Privada", "Limite_sup_Privada"
        ]
        suffix = "Estruturado_Regioes_UF"
    
    # Formatação dos cabeçalhos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, cabecalho in enumerate(cabecalhos, 1):
        cell = ws_consolidado.cell(row=1, column=col_idx)
        cell.value = cabecalho
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    
    linha_atual = 2
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Mapeamento para Aba 2
    regioes_validas = ["Brasil", "Norte", "Nordeste", "Sudeste", "Sul", "Centro-Oeste"]
    estados_por_regiao = {
        "Norte": ["Rondônia", "Acre", "Amazonas", "Roraima", "Pará", "Amapá", "Tocantins"],
        "Nordeste": ["Maranhão", "Piauí", "Ceará", "Rio Grande do Norte", "Paraíba", "Pernambuco", "Alagoas", "Sergipe", "Bahia"],
        "Sudeste": ["Minas Gerais", "Espírito Santo", "Rio de Janeiro", "São Paulo"],
        "Sul": ["Paraná", "Santa Catarina", "Rio Grande do Sul"],
        "Centro-Oeste": ["Mato Grosso do Sul", "Mato Grosso", "Goiás", "Distrito Federal"]
    }

    for arquivo_name, tabela_origem, nome_metrica in arquivos:
        caminho = os.path.join(diretorio_origem, arquivo_name)
        if not os.path.exists(caminho):
            print(f"✗ Arquivo não encontrado: {arquivo_name}")
            continue

        try:
            wb = openpyxl.load_workbook(caminho, data_only=True)
            ws = wb.worksheets[aba_indice]
            dados = list(ws.iter_rows(values_only=True))
            
            # Localizar linha de início (Brasil)
            inicio_dados = 0
            for idx, row in enumerate(dados):
                if row and any("Brasil" in str(cell) for cell in row if cell is not None):
                    inicio_dados = idx
                    break
            
            faixa_atual = None
            regiao_atual = "Brasil"
            
            for row_idx in range(inicio_dados, len(dados)):
                row = dados[row_idx]
                if not row or row[0] is None: continue
                
                item = str(row[0]).strip()
                if item.startswith("Fonte:") or not item: continue
                
                # Lógica para Aba 1 (Idades)
                if aba_indice == 0:
                    if "anos" in item:
                        faixa_atual = item
                        continue
                    if item in regioes_validas and faixa_atual:
                        nova_linha = [tabela_origem, nome_metrica, faixa_atual, item]
                        # Pega colunas 1 a 15
                        for i in range(1, 16):
                            val = row[i] if i < len(row) else None
                            nova_linha.append(round(val, 2) if isinstance(val, (int, float)) else val)
                    else: continue

                # Lógica para Aba 2 (Regiões/UF)
                else:
                    if item in regioes_validas: regiao_atual = item
                    is_estado = any(item == est for lista in estados_por_regiao.values() for est in lista)
                    
                    if item in regioes_validas or is_estado:
                        reg = item if item in regioes_validas else regiao_atual
                        uf = item if is_estado else None
                        nova_linha = [tabela_origem, nome_metrica, reg, uf]
                        for i in range(1, 16):
                            val = row[i] if i < len(row) else None
                            nova_linha.append(round(val, 2) if isinstance(val, (int, float)) else val)
                    else: continue

                # Escrever no Excel Consolidado
                for col_idx, valor in enumerate(nova_linha[:len(cabecalhos)], 1):
                    cell = ws_consolidado.cell(row=linha_atual, column=col_idx)
                    cell.value = valor
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
                    if col_idx > 4 and isinstance(valor, (int, float)):
                        cell.number_format = '0.00'
                linha_atual += 1

            wb.close()
            print(f"✓ {arquivo_name} processado.")
            
        except Exception as e:
            print(f"✗ Erro em {arquivo_name}: {e}")

    # Ajuste de larguras e salvamento
    for col_idx in range(1, len(cabecalhos) + 1):
        ws_consolidado.column_dimensions[get_column_letter(col_idx)].width = 25 if col_idx <= 4 else 12

    output_path = os.path.join(diretorio_origem, f"Tabelas_12_Consolidadas_{suffix}.xlsx")
    wb_consolidado.save(output_path)
    print(f"\n★ Concluído! Arquivo salvo em:\n{output_path}")

if __name__ == "__main__":
    import sys
    # Se rodar sem argumentos, processa a primeira aba (0). 
    # Para a segunda aba, rode: python nome_do_script.py 1
    indice = int(sys.argv[1]) if len(sys.argv) > 1 else 0
    consolidar_tabelas(aba_indice=indice)