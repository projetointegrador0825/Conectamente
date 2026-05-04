#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Script consolidado para a SEGUNDA ABA (Grandes Regiões e Unidades da Federação)
Caminho e lógica de diretórios atualizados para ambiente Windows.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def consolidar_tabelas_regioes_uf(diretorio_origem=None):
    """
    Consolida as Tabelas 12.1 a 12.8 (segunda aba) em um Excel estruturado.
    """
    
    # Atualizado conforme especificação e imagem image_49c601.png
    if diretorio_origem is None:
        diretorio_origem = r"C:\Users\caiol\Desktop\PROJETO INTEGRADOR 2 ETAPA\projeto-integrador-grupo26\tarefa_caio\dados"
    
    arquivos = [
        ("Tabelas 12.1.xlsx", "12.1.2", "NÃO_TEM_AMIGO_PROXIMO"),
        ("Tabelas 12.2.xlsx", "12.2.2", "MUITO_PREOCUPADO"),
        ("Tabelas 12.3.xlsx", "12.3.2", "TRISTE"),
        ("Tabelas 12.4.xlsx", "12.4.2", "SOZINHO"),
        ("Tabelas 12.5.xlsx", "12.5.2", "PROBLEMA_CONCENTRACAO"),
        ("Tabelas 12.6.xlsx", "12.6.2", "PROBLEMA_INSONIA"),
        ("Tabelas 12.7.xlsx", "12.7.2", "SENTIMENTO_INUTILIDADE"),
        ("Tabelas 12.8.xlsx", "12.8.2", "PENSAMENTO_MORTE"),
    ]
    
    wb_consolidado = Workbook()
    ws_consolidado = wb_consolidado.active
    ws_consolidado.title = "Consolidado_Regioes_UF"
    
    cabecalhos = [
        "TABELA_ORIGEM", "NOME_METRICA", "Regiao", "Unidade_Federacao", "rc",
        "Limite_inferior", "Limite_superior",
        "rc_Homem", "Limite_inf_Homem", "Limite_sup_Homem",
        "rc_Mulher", "Limite_inf_Mulher", "Limite_sup_Mulher",
        "rc_Publica", "Limite_inf_Publica", "Limite_sup_Publica",
        "rc_Privada", "Limite_inf_Privada", "Limite_sup_Privada"
    ]
    
    # Estilização
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    center_align = Alignment(wrap_text=True, vertical="center", horizontal="center")
    
    for col_idx, cabecalho in enumerate(cabecalhos, 1):
        cell = ws_consolidado.cell(row=1, column=col_idx)
        cell.value = cabecalho
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    
    linha_atual = 2
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    regioes_validas = ["Brasil", "Norte", "Nordeste", "Sudeste", "Sul", "Centro-Oeste"]
    estados_por_regiao = {
        "Norte": ["Rondônia", "Acre", "Amazonas", "Roraima", "Pará", "Amapá", "Tocantins"],
        "Nordeste": ["Maranhão", "Piauí", "Ceará", "Rio Grande do Norte", "Paraíba", "Pernambuco", "Alagoas", "Sergipe", "Bahia"],
        "Sudeste": ["Minas Gerais", "Espírito Santo", "Rio de Janeiro", "São Paulo"],
        "Sul": ["Paraná", "Santa Catarina", "Rio Grande do Sul"],
        "Centro-Oeste": ["Mato Grosso do Sul", "Mato Grosso", "Goiás", "Distrito Federal"]
    }
    
    for arquivo_name, tabela_origem, nome_metrica in arquivos:
        caminho = os.path.join(diretorio_origem, arquivo_name)
        
        if not os.path.exists(caminho):
            print(f"✗ Arquivo não encontrado: {arquivo_name}")
            continue

        try:
            wb = openpyxl.load_workbook(caminho, data_only=True)
            
            if len(wb.worksheets) < 2:
                print(f"⚠ {arquivo_name} não possui a segunda aba.")
                wb.close()
                continue
                
            ws = wb.worksheets[1] # Segunda aba (índice 1)
            dados = list(ws.iter_rows(values_only=True))
            
            inicio_dados = 0
            for idx, row in enumerate(dados):
                if row and any("Brasil" in str(cell) for cell in row if cell is not None):
                    inicio_dados = idx
                    break
            
            regiao_atual = "Brasil"
            
            for row_idx in range(inicio_dados, len(dados)):
                row = dados[row_idx]
                if not row or row[0] is None: continue
                
                item = str(row[0]).strip()
                if item.startswith("Fonte:") or not item: continue
                
                # Atualiza região atual se encontrar uma nova
                if item in regioes_validas:
                    regiao_atual = item
                
                # Verifica se é um item válido (Região ou Estado)
                is_estado = any(item == estado for lista in estados_por_regiao.values() for estado in lista)
                
                if item in regioes_validas or is_estado:
                    regiao = item if item in regioes_validas else regiao_atual
                    unidade_fed = item if is_estado else None
                    
                    nova_linha = [tabela_origem, nome_metrica, regiao, unidade_fed]
                    
                    # Coleta as colunas numéricas
                    for i in range(1, 16):
                        valor = row[i] if i < len(row) else None
                        nova_linha.append(round(valor, 2) if isinstance(valor, (int, float)) else valor)
                    
                    while len(nova_linha) < len(cabecalhos):
                        nova_linha.append(None)

                    for col_idx, valor in enumerate(nova_linha[:len(cabecalhos)], 1):
                        cell = ws_consolidado.cell(row=linha_atual, column=col_idx)
                        cell.value = valor
                        cell.border = thin_border
                        cell.alignment = Alignment(wrap_text=True, vertical="center")
                        if col_idx > 4 and isinstance(valor, (int, float)):
                            cell.number_format = '0.00'
                    
                    linha_atual += 1
            
            wb.close()
            print(f"✓ {arquivo_name} (Aba 2) processado.")
            
        except Exception as e:
            print(f"✗ Erro em {arquivo_name}: {e}")

    # Ajuste de larguras das colunas
    for col_idx in range(1, len(cabecalhos) + 1):
        ws_consolidado.column_dimensions[get_column_letter(col_idx)].width = 25 if col_idx <= 4 else 12

    # Salva o resultado final no diretório de dados identificado na image_49c601.png
    output_path = os.path.join(diretorio_origem, "Tabelas_12_Consolidadas_Regioes_UF.xlsx")
    wb_consolidado.save(output_path)
    print(f"\n★ Concluído! Arquivo salvo em:\n{output_path}")

if __name__ == "__main__":
    consolidar_tabelas_regioes_uf()