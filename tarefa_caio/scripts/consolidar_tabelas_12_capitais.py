#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Script consolidado para a TERCEIRA ABA (Municípios das Capitais)
Caminho atualizado conforme estrutura de diretórios do Windows.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def consolidar_tabelas_municipios(diretorio_origem=None):
    """
    Consolida as Tabelas 12.1 a 12.8 (terceira aba) em um Excel estruturado.
    """
    
    # Atualizado conforme sua especificação
    if diretorio_origem is None:
        diretorio_origem = r"C:\Users\caiol\Desktop\PROJETO INTEGRADOR 2 ETAPA\projeto-integrador-grupo26\tarefa_caio\dados"
    
    arquivos = [
        ("Tabelas 12.1.xlsx", "12.1.3", "NÃO_TEM_AMIGO_PROXIMO"),
        ("Tabelas 12.2.xlsx", "12.2.3", "MUITO_PREOCUPADO"),
        ("Tabelas 12.3.xlsx", "12.3.3", "TRISTE"),
        ("Tabelas 12.4.xlsx", "12.4.3", "SOZINHO"),
        ("Tabelas 12.5.xlsx", "12.5.3", "PROBLEMA_CONCENTRACAO"),
        ("Tabelas 12.6.xlsx", "12.6.3", "PROBLEMA_INSONIA"),
        ("Tabelas 12.7.xlsx", "12.7.3", "SENTIMENTO_INUTILIDADE"),
        ("Tabelas 12.8.xlsx", "12.8.3", "PENSAMENTO_MORTE"),
    ]
    
    wb_consolidado = Workbook()
    ws_consolidado = wb_consolidado.active
    ws_consolidado.title = "Consolidado_Municipios"
    
    cabecalhos = [
        "TABELA_ORIGEM", "NOME_METRICA", "Municipio", "rc",
        "Limite_inferior", "Limite_superior",
        "rc_Homem", "Limite_inf_Homem", "Limite_sup_Homem",
        "rc_Mulher", "Limite_inf_Mulher", "Limite_sup_Mulher",
        "rc_Publica", "Limite_inf_Publica", "Limite_sup_Publica",
        "rc_Privada", "Limite_inf_Privada", "Limite_sup_Privada"
    ]
    
    # Estilização
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    center_align = Alignment(wrap_text=True, vertical="center", horizontal="center")
    
    for col_idx, cabecalho in enumerate(cabecalhos, 1):
        cell = ws_consolidado.cell(row=1, column=col_idx)
        cell.value = cabecalho
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    
    linha_atual = 2
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for arquivo_name, tabela_origem, nome_metrica in arquivos:
        caminho = os.path.join(diretorio_origem, arquivo_name)
        
        if not os.path.exists(caminho):
            print(f"✗ Arquivo não encontrado: {arquivo_name}")
            continue

        try:
            wb = openpyxl.load_workbook(caminho, data_only=True)
            
            if len(wb.worksheets) < 3:
                print(f"⚠ {arquivo_name} não possui a terceira aba.")
                wb.close()
                continue
                
            ws = wb.worksheets[2] 
            dados = list(ws.iter_rows(values_only=True))
            
            # Localizar linha de início (Total)
            inicio_dados = 0
            for idx, row in enumerate(dados):
                if row and any("Total" in str(cell) for cell in row if cell is not None):
                    inicio_dados = idx
                    break
            
            for row_idx in range(inicio_dados, len(dados)):
                row = dados[row_idx]
                if not row or row[0] is None: continue
                
                municipio = str(row[0]).strip()
                if municipio.startswith("Fonte:") or not municipio: continue
                
                nova_linha = [tabela_origem, nome_metrica, municipio]
                
                # Coleta colunas de dados
                for i in range(1, 16):
                    valor = row[i] if i < len(row) else None
                    nova_linha.append(round(valor, 2) if isinstance(valor, (int, float)) else valor)
                
                # Preencher colunas vazias para manter padrão
                while len(nova_linha) < len(cabecalhos):
                    nova_linha.append(None)

                for col_idx, valor in enumerate(nova_linha[:len(cabecalhos)], 1):
                    cell = ws_consolidado.cell(row=linha_atual, column=col_idx)
                    cell.value = valor
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
                    if col_idx > 3 and isinstance(valor, (int, float)):
                        cell.number_format = '0.00'
                
                linha_atual += 1
            
            wb.close()
            print(f"✓ {arquivo_name} processado.")
            
        except Exception as e:
            print(f"✗ Erro em {arquivo_name}: {e}")

    # Ajuste de larguras
    for col_idx in range(1, len(cabecalhos) + 1):
        ws_consolidado.column_dimensions[get_column_letter(col_idx)].width = 25 if col_idx <= 3 else 12

    # Salva na mesma pasta de origem para facilitar o acesso
    output_path = os.path.join(diretorio_origem, "Tabelas_12_Consolidadas_Municipios.xlsx")
    
    wb_consolidado.save(output_path)
    print(f"\n★ Concluído! Arquivo salvo em:\n{output_path}")

if __name__ == "__main__":
    consolidar_tabelas_municipios()